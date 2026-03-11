import asyncio
import json
import os
import re
import tempfile
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.exceptions import TelegramBadRequest
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import CallbackQuery, FSInputFile, InlineKeyboardButton, InlineKeyboardMarkup, Message
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

CONFIG_PATH = os.path.join(os.path.dirname(__file__), "config.json")
SERVICES_SHEET_NAME = "services"
SERVICES_HEADERS = ["Категория", "Модель", "Услуга", "Цена"]


def load_config() -> Dict[str, Any]:
    if not os.path.exists(CONFIG_PATH):
        raise FileNotFoundError(f"Config file not found: {CONFIG_PATH}")
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_config(cfg: Dict[str, Any]) -> None:
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def get_cfg() -> Dict[str, Any]:
    global config
    config = load_config()
    return config


config = load_config()

REPAIR_PROBLEM_KEYS = [
    ("screen", "Экран"),
    ("battery", "Аккумулятор"),
    ("no_power", "Не включается"),
    ("cleaning", "Чистка"),
    ("upgrade", "Увеличение"),
    ("keyboard", "Клавиатура"),
    ("firmware", "Прошивка"),
    ("water_damage", "После влаги"),
    ("other", "Другое"),
]

ADMIN_FIELD_LABELS = {
    "welcome_text": "Приветствие",
    "menu_text": "Текст перед выбором услуги",
    "hydro_text": "Текст «Гидрогелевая плёнка»",
    "price_text": "Текст «Стоимость защиты и аксессуаров»",
    "warranty_text": "Текст «Хочу обратиться по гарантии»",
    "device_in_service_prompt": "Подсказка «Моё устройство в сервисе»",
    "device_in_service_phone_question": "Вопрос про доступность по номеру",
    "ask_contact_text": "Запрос контакта (общий)",
    "offices": "Офисы (JSON-массив строк)",
    "office_addresses": "Адреса офисов (JSON-массив строк)",
    "repair_time_slots": "Слоты времени (JSON-массив строк)",
    "repair_confirmation_template": "Шаблон подтверждения записи",
    "excel_catalog_download": "Скачать Excel-каталог",
    "excel_catalog_upload": "Загрузить Excel-каталог",
}

ADMIN_SECTIONS = {
    "texts": "Тексты",
    "json": "Json",
    "templates": "Шаблоны",
}

ADMIN_FIELDS_BY_SECTION = {
    "texts": [
        "welcome_text",
        "menu_text",
        "hydro_text",
        "price_text",
        "warranty_text",
        "device_in_service_prompt",
        "device_in_service_phone_question",
        "ask_contact_text",
    ],
    "json": [
        "offices",
        "office_addresses",
        "repair_time_slots",
        "excel_catalog_download",
        "excel_catalog_upload",
    ],
    "templates": [
        "repair_confirmation_template",
    ],
}

ADMIN_JSON_FIELDS = {
    "offices",
    "office_addresses",
    "repair_time_slots",
}


class UserStates(StatesGroup):
    waiting_device_office = State()
    waiting_device_info = State()
    waiting_contact_for_device = State()
    repair_choose_category = State()
    repair_choose_model = State()
    repair_choose_problem = State()
    repair_other_problem = State()
    repair_description = State()
    repair_choose_office = State()
    repair_choose_day = State()
    repair_enter_date = State()
    repair_choose_time = State()
    repair_ask_name = State()
    repair_ask_contact = State()


class AdminEditStates(StatesGroup):
    waiting_new_value = State()
    waiting_excel_upload = State()


def back_keyboard(callback_data: str, text: str = "⬅️ Назад") -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[[InlineKeyboardButton(text=text, callback_data=callback_data)]]
    )


def get_admin_ids() -> List[int]:
    raw = os.getenv("ADMIN_IDS", "").strip()
    if not raw:
        raise RuntimeError("ADMIN_IDS is not set in environment")
    result: List[int] = []
    for part in raw.split(","):
        value = part.strip()
        if not value:
            continue
        try:
            result.append(int(value))
        except ValueError:
            print(f"[WARN] ADMIN_IDS contains invalid value: {value}")
    if not result:
        raise RuntimeError("ADMIN_IDS does not contain valid IDs")
    return result


def is_admin(user_id: int) -> bool:
    return user_id in get_admin_ids()


def chunks(lst: List[Any], n: int) -> List[List[Any]]:
    return [lst[i : i + n] for i in range(0, len(lst), n)]


def normalize_text(value: str) -> str:
    # Allow admins to enter \n in config and get real line breaks in messages.
    return value.replace("\\n", "\n")


def cfg_text(key: str, default: str) -> str:
    get_cfg()
    value = config.get(key, default)
    if not isinstance(value, str):
        return default
    return normalize_text(value)


def get_services_xlsx_path() -> str:
    return os.getenv(
        "SERVICES_XLSX_PATH", os.path.join(os.path.dirname(__file__), "services.xlsx")
    )


def repair_problem_category_key(category_key: str) -> str:
    return f"repair_problems_{category_key}"


def build_legacy_service_rows() -> List[List[str]]:
    get_cfg()
    rows: List[List[str]] = []
    categories = config.get("repair_categories", [])
    models_by_category = config.get("repair_models_by_category", {})

    for item in categories:
        if not isinstance(item, dict):
            continue
        category_key = str(item.get("key", "")).strip()
        category_title = str(item.get("title", category_key)).strip()
        if not category_key or not category_title:
            continue

        if category_key == "laptops":
            models = config.get("repair_laptop_models", [])
        else:
            models = models_by_category.get(category_key, [])

        raw_services = config.get(repair_problem_category_key(category_key), [])
        if not raw_services:
            raw_services = config.get("repair_problems_by_category", {}).get(category_key, [])

        for model in models:
            model_title = str(model).strip()
            if not model_title:
                continue
            for service in raw_services:
                if isinstance(service, dict):
                    service_title = str(service.get("title", "")).strip()
                    price = str(service.get("price", "")).strip()
                else:
                    service_title = str(service).strip()
                    price = ""
                if service_title:
                    rows.append([category_title, model_title, service_title, price])

    return rows


def ensure_services_workbook() -> str:
    path = get_services_xlsx_path()
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    if os.path.exists(path):
        return path

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SERVICES_SHEET_NAME
    sheet.append(SERVICES_HEADERS)
    for row in build_legacy_service_rows():
        sheet.append(row)
    workbook.save(path)
    return path


def validate_services_workbook(path: str) -> List[Dict[str, str]]:
    workbook = load_workbook(path, data_only=True)
    if SERVICES_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"В Excel нет листа '{SERVICES_SHEET_NAME}'.")

    sheet = workbook[SERVICES_SHEET_NAME]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    missing = [header for header in SERVICES_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Не найдены обязательные столбцы: {', '.join(missing)}")

    indexes = {header: headers.index(header) for header in SERVICES_HEADERS}
    rows: List[Dict[str, str]] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = ["" if value is None else str(value).strip() for value in row]
        if not any(values):
            continue

        record = {}
        for header in SERVICES_HEADERS:
            index = indexes[header]
            value = values[index] if index < len(values) else ""
            record[header] = value

        if not all(record.values()):
            raise ValueError(
                f"Строка {row_idx} заполнена не полностью. Нужны значения в колонках: "
                f"{', '.join(SERVICES_HEADERS)}"
            )
        rows.append(record)

    if not rows:
        raise ValueError("Excel-файл пустой. Добавьте хотя бы одну строку услуги.")

    return rows


def load_services_catalog() -> Dict[str, Any]:
    path = ensure_services_workbook()
    rows = validate_services_workbook(path)

    categories: List[str] = []
    models_by_category: Dict[str, List[str]] = {}
    services_by_pair: Dict[str, List[Dict[str, str]]] = {}
    seen_services: Dict[tuple[str, str, str], Dict[str, str]] = {}
    order: List[tuple[str, str, str]] = []

    for row in rows:
        category = row["Категория"]
        model = row["Модель"]
        service = row["Услуга"]
        price = row["Цена"]

        if category not in categories:
            categories.append(category)
        models_by_category.setdefault(category, [])
        if model not in models_by_category[category]:
            models_by_category[category].append(model)

        key = (category, model, service)
        if key not in seen_services:
            order.append(key)
        seen_services[key] = {
            "title": service,
            "price": price,
        }

    for category, model, service in order:
        pair_key = f"{category}|||{model}"
        services_by_pair.setdefault(pair_key, []).append(seen_services[(category, model, service)])

    return {
        "categories": categories,
        "models_by_category": models_by_category,
        "services_by_pair": services_by_pair,
        "path": path,
    }


def main_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="Ремонт / диагностика", callback_data="menu_repair")],
            [InlineKeyboardButton(text="Гидрогелевая плёнка", callback_data="menu_hydro")],
            [InlineKeyboardButton(text="Стоимость защиты и аксессуаров", callback_data="menu_price")],
            [InlineKeyboardButton(text="Моё устройство в сервисе", callback_data="menu_device_service")],
            [InlineKeyboardButton(text="Хочу обратиться по гарантии", callback_data="menu_warranty")],
        ]
    )


def yes_no_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="Да", callback_data="yes_available"),
                InlineKeyboardButton(text="Нет", callback_data="no_available"),
            ],
            [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_device_info")],
        ]
    )


def admin_menu_keyboard() -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton(text=label, callback_data=f"admin_section_{key}")]
        for key, label in ADMIN_SECTIONS.items()
    ]
    return InlineKeyboardMarkup(inline_keyboard=rows)


def admin_section_keyboard(section_key: str) -> InlineKeyboardMarkup:
    field_keys = ADMIN_FIELDS_BY_SECTION.get(section_key, [])
    rows = [
        [InlineKeyboardButton(text=ADMIN_FIELD_LABELS[field], callback_data=f"admin_edit_{field}")]
        for field in field_keys
        if field in ADMIN_FIELD_LABELS
    ]
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="admin_root")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def admin_problem_categories_keyboard() -> InlineKeyboardMarkup:
    get_cfg()
    rows: List[List[InlineKeyboardButton]] = []
    for item in config.get("repair_categories", []):
        key = item.get("key")
        title = item.get("title", key)
        if not key:
            continue
        rows.append(
            [
                InlineKeyboardButton(
                    text=title,
                    callback_data=f"admin_edit_problem_cat_{key}",
                )
            ]
        )
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="admin_section_json")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def repair_categories_keyboard(categories: List[str]) -> InlineKeyboardMarkup:
    rows: List[List[InlineKeyboardButton]] = []
    indexed_categories = list(enumerate(categories))
    for block in chunks(indexed_categories, 3):
        rows.append(
            [
                InlineKeyboardButton(text=title, callback_data=f"repair_cat_{idx}")
                for idx, title in block
            ]
        )
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="back_main_menu")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def repair_models_keyboard(models: List[str], back_callback: str = "back_repair_categories") -> InlineKeyboardMarkup:
    rows: List[List[InlineKeyboardButton]] = []
    indexed_models = list(enumerate(models))
    for block in chunks(indexed_models, 3):
        rows.append(
            [InlineKeyboardButton(text=name, callback_data=f"repair_model_{idx}") for idx, name in block]
        )
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data=back_callback)])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def repair_services_keyboard(
    services: List[Dict[str, str]], back_callback: str
) -> InlineKeyboardMarkup:
    rows: List[List[InlineKeyboardButton]] = []
    indexed_services = list(enumerate(services))
    for block in chunks(indexed_services, 3):
        rows.append(
            [
                InlineKeyboardButton(
                    text=item.get("title", "Услуга"),
                    callback_data=f"repair_service_{idx}",
                )
                for idx, item in block
            ]
        )
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data=back_callback)])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def offices_keyboard(prefix: str, back_callback: Optional[str] = None) -> InlineKeyboardMarkup:
    get_cfg()
    rows = [
        [InlineKeyboardButton(text=office, callback_data=f"{prefix}_{idx}")]
        for idx, office in enumerate(config.get("offices", []))
    ]
    if back_callback:
        rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data=back_callback)])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def repair_day_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="Завтра", callback_data="repair_day_tomorrow")],
            [InlineKeyboardButton(text="Другой день", callback_data="repair_day_other")],
            [InlineKeyboardButton(text="⬅️ Назад", callback_data="back_repair_office")],
        ]
    )


def repair_time_keyboard() -> InlineKeyboardMarkup:
    get_cfg()
    slots = config.get("repair_time_slots", [])
    any_time = config.get("repair_any_time_text", "Приду в течение дня")
    rows: List[List[InlineKeyboardButton]] = []
    for block in chunks(list(enumerate(slots)), 3):
        rows.append([InlineKeyboardButton(text=slot, callback_data=f"repair_time_{idx}") for idx, slot in block])
    rows.append([InlineKeyboardButton(text=any_time, callback_data="repair_time_any")])
    rows.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="back_repair_day")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


async def send_to_admin(bot: Bot, text: str) -> bool:
    delivered = False
    for admin_id in get_admin_ids():
        try:
            await bot.send_message(chat_id=admin_id, text=text)
            delivered = True
        except TelegramBadRequest as e:
            print(f"[WARN] Failed to send to admin {admin_id}: {e}")
    return delivered


def is_phone_number(value: str) -> bool:
    return value.isdigit() and len(value) == 11 and value.startswith("7")


def is_order_number(value: str) -> bool:
    return bool(re.fullmatch(r"\d{3}-\d{5}", value))


def parse_date_ru(value: str) -> Optional[datetime]:
    try:
        return datetime.strptime(value, "%d.%m.%Y")
    except ValueError:
        return None


def office_address_by_idx(idx: int) -> str:
    get_cfg()
    addresses = config.get("office_addresses", [])
    if 0 <= idx < len(addresses):
        return addresses[idx]
    offices = config.get("offices", [])
    return offices[idx] if 0 <= idx < len(offices) else "Адрес уточнит администратор"


async def start_handler(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(cfg_text("welcome_text", "Здравствуйте!"), reply_markup=main_menu_keyboard())


async def menu_repair_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    catalog = load_services_catalog()
    await state.set_state(UserStates.repair_choose_category)
    await state.update_data(repair_catalog=catalog)
    await callback.message.edit_text(
        "Выберите категорию устройства:",
        reply_markup=repair_categories_keyboard(catalog.get("categories", [])),
    )


async def repair_category_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    idx_raw = callback.data.replace("repair_cat_", "", 1)
    if not idx_raw.isdigit():
        return
    idx = int(idx_raw)
    data = await state.get_data()
    catalog = data.get("repair_catalog") or load_services_catalog()
    categories = catalog.get("categories", [])
    if idx < 0 or idx >= len(categories):
        return

    category = categories[idx]
    models = catalog.get("models_by_category", {}).get(category, [])
    await state.update_data(
        repair_category=category,
        repair_models=models,
    )
    await state.set_state(UserStates.repair_choose_model)
    await callback.message.edit_text(
        "Выберите модель:",
        reply_markup=repair_models_keyboard(models, "back_repair_categories"),
    )


async def repair_model_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    idx_raw = callback.data.replace("repair_model_", "", 1)
    if not idx_raw.isdigit():
        return
    idx = int(idx_raw)
    data = await state.get_data()
    models = data.get("repair_models", [])
    if idx < 0 or idx >= len(models):
        return

    category = data.get("repair_category", "")
    model = models[idx]
    catalog = data.get("repair_catalog") or load_services_catalog()
    pair_key = f"{category}|||{model}"
    services = catalog.get("services_by_pair", {}).get(pair_key, [])

    await state.update_data(
        repair_device=model,
        repair_services=services,
    )
    await state.set_state(UserStates.repair_choose_problem)
    await callback.message.edit_text(
        "Выберите услугу:",
        reply_markup=repair_services_keyboard(services, "back_repair_models"),
    )


async def repair_problem_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    idx_raw = callback.data.replace("repair_service_", "", 1)
    if not idx_raw.isdigit():
        return
    idx = int(idx_raw)
    data = await state.get_data()
    services = data.get("repair_services", [])
    if idx < 0 or idx >= len(services):
        return
    selected = services[idx]
    title = selected.get("title", "Услуга")
    price = selected.get("price", "")
    text = f"Услуга: {title}"
    if price:
        text = f"{text}\n\n💵 Стоимость: {price}"
    else:
        text = f"{text}\n\n💵 Стоимость уточним после диагностики."

    await state.update_data(repair_service_title=title, repair_service_price=price)
    await state.set_state(UserStates.repair_description)
    await callback.message.edit_text(
        f"{text}\n\n{cfg_text('repair_description_prompt', 'Опишите проблему подробнее.')}",
        reply_markup=back_keyboard("back_repair_problems"),
    )


async def repair_other_problem_handler(message: Message, state: FSMContext) -> None:
    await repair_description_handler(message, state)


async def repair_description_handler(message: Message, state: FSMContext) -> None:
    await state.update_data(repair_description=(message.text or "").strip())
    await state.set_state(UserStates.repair_choose_office)
    await message.answer(
        cfg_text("repair_choose_office_text", "Выберите офис:"),
        reply_markup=offices_keyboard("repair_office", "back_repair_description"),
    )


async def repair_office_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    idx_raw = callback.data.replace("repair_office_", "", 1)
    if not idx_raw.isdigit():
        return
    idx = int(idx_raw)
    offices = config.get("offices", [])
    if idx < 0 or idx >= len(offices):
        return
    await state.update_data(repair_office_idx=idx, repair_office_name=offices[idx])
    await state.set_state(UserStates.repair_choose_day)
    await callback.message.edit_text(
        cfg_text("repair_choose_day_text", "Выберите день записи:"),
        reply_markup=repair_day_keyboard(),
    )


async def repair_day_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    if callback.data == "repair_day_tomorrow":
        date_str = (datetime.now() + timedelta(days=1)).strftime("%d.%m.%Y")
        await state.update_data(repair_date=date_str)
        await state.set_state(UserStates.repair_choose_time)
        await callback.message.edit_text(
            cfg_text("repair_choose_time_text", "Выберите время:"),
            reply_markup=repair_time_keyboard(),
        )
        return
    await state.set_state(UserStates.repair_enter_date)
    await callback.message.edit_text(
        cfg_text("repair_enter_date_text", "Введите дату ДД.ММ.ГГГГ")
    )


async def repair_date_input_handler(message: Message, state: FSMContext) -> None:
    date_raw = (message.text or "").strip()
    dt = parse_date_ru(date_raw)
    if dt is None:
        await message.answer("Неверный формат даты. Используйте ДД.ММ.ГГГГ.")
        return
    await state.update_data(repair_date=date_raw)
    await state.set_state(UserStates.repair_choose_time)
    await message.answer(
        cfg_text("repair_choose_time_text", "Выберите время:"),
        reply_markup=repair_time_keyboard(),
    )


async def repair_time_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    slots = config.get("repair_time_slots", [])
    any_time = cfg_text("repair_any_time_text", "Приду в течение дня")
    if callback.data == "repair_time_any":
        picked = any_time
    else:
        idx_raw = callback.data.replace("repair_time_", "", 1)
        if not idx_raw.isdigit():
            return
        idx = int(idx_raw)
        if idx < 0 or idx >= len(slots):
            return
        picked = slots[idx]
    await state.update_data(repair_time=picked)
    await state.set_state(UserStates.repair_ask_name)
    await callback.message.edit_text(
        cfg_text("repair_ask_name_text", "Как вас зовут?"),
        reply_markup=back_keyboard("back_repair_time"),
    )


async def repair_name_handler(message: Message, state: FSMContext) -> None:
    await state.update_data(repair_name=(message.text or "").strip())
    await state.set_state(UserStates.repair_ask_contact)
    await message.answer(
        cfg_text("repair_ask_contact_text", "Укажите телефон или @username"),
        reply_markup=back_keyboard("back_repair_name"),
    )


async def repair_contact_handler(message: Message, state: FSMContext) -> None:
    get_cfg()
    contact = (message.text or "").strip()
    data = await state.get_data()
    date_val = data.get("repair_date", "дата не указана")
    time_val = data.get("repair_time", "время не указано")
    device_val = data.get("repair_device", "устройство")
    category_val = data.get("repair_category", "-")
    service_val = data.get("repair_service_title", "услуга")
    service_price = data.get("repair_service_price", "")
    office_idx = data.get("repair_office_idx", 0)
    office_name = data.get("repair_office_name", "Офис")
    office_address = office_address_by_idx(office_idx)

    confirmation_template = cfg_text(
        "repair_confirmation_template",
        "Ок, записал вас на {date} в {time} на {problem} для устройства {device}.\n\nЖдём вас по адресу: {office}",
    )
    user_text = confirmation_template.format(
        date=date_val,
        time=time_val,
        problem=service_val,
        device=device_val,
        office=office_address,
    )

    admin_text = (
        f"{cfg_text('repair_send_to_admin_header', '🛠 Новая запись на ремонт / диагностику')}\n\n"
        f"Категория: {category_val}\n"
        f"Модель/устройство: {device_val}\n"
        f"Услуга: {service_val}\n"
        f"Цена: {service_price or 'не указана'}\n"
        f"Описание: {data.get('repair_description', '-')}\n\n"
        f"Офис: {office_name}\n"
        f"Адрес: {office_address}\n"
        f"Дата: {date_val}\n"
        f"Слот: {time_val}\n\n"
        f"Клиент: {data.get('repair_name', '-')}\n"
        f"Контакт: {contact}\n"
        f"TG user: {message.from_user.full_name} (@{message.from_user.username})\n"
        f"ID: {message.from_user.id}"
    )
    delivered = await send_to_admin(message.bot, admin_text)
    await state.clear()
    await message.answer(
        f"{user_text}\n\n"
        + (
            "Контакт для связи получили. Главное меню:"
            if delivered
            else "Контакт получили, но не смогли отправить заявку администратору. Проверьте список ADMIN_IDS."
        ),
        reply_markup=main_menu_keyboard(),
    )


async def menu_hydro_handler(callback: CallbackQuery) -> None:
    await callback.answer()
    await callback.message.edit_text(
        cfg_text("hydro_text", "Информация о гидрогелевой плёнке."),
        reply_markup=main_menu_keyboard(),
    )


async def menu_price_handler(callback: CallbackQuery) -> None:
    await callback.answer()
    await callback.message.edit_text(
        cfg_text("price_text", "Информация о стоимости защиты и аксессуаров."),
        reply_markup=main_menu_keyboard(),
    )


async def menu_warranty_handler(callback: CallbackQuery) -> None:
    await callback.answer()
    await callback.message.edit_text(
        cfg_text("warranty_text", "Информация по гарантии."),
        reply_markup=main_menu_keyboard(),
    )


async def menu_device_service_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    await state.set_state(UserStates.waiting_device_office)
    await callback.message.edit_text(
        "Выберите офис, в который обращались:",
        reply_markup=offices_keyboard("device_office", "back_main_menu"),
    )


async def device_office_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    idx_raw = callback.data.replace("device_office_", "", 1)
    if not idx_raw.isdigit():
        return
    idx = int(idx_raw)
    offices = config.get("offices", [])
    office_name = offices[idx] if 0 <= idx < len(offices) else "Неизвестный офис"
    await state.update_data(chosen_office=office_name)
    await state.set_state(UserStates.waiting_device_info)
    await callback.message.edit_text(
        cfg_text("device_in_service_prompt", "Напишите номер заказа или телефон, указанный в заказе."),
        reply_markup=back_keyboard("back_device_office"),
    )


async def handle_device_info(message: Message, state: FSMContext) -> None:
    value = (message.text or "").strip()
    await state.update_data(device_raw_input=value)
    if is_phone_number(value):
        await state.update_data(device_phone=value)
        await message.answer(
            cfg_text("device_in_service_phone_question", "Вы сейчас доступны по этому номеру телефона?"),
            reply_markup=yes_no_keyboard(),
        )
        return
    await state.update_data(device_order=value if is_order_number(value) else None)
    await state.set_state(UserStates.waiting_contact_for_device)
    await message.answer(
        cfg_text("ask_contact_text", "Пожалуйста, отправьте телефон или @username"),
        reply_markup=back_keyboard("back_device_info"),
    )


async def handle_device_contact(message: Message, state: FSMContext) -> None:
    contact = (message.text or "").strip()
    data = await state.get_data()
    admin_text = (
        "📦 Запрос по устройству в сервисе\n\n"
        f"Офис: {data.get('chosen_office', 'Не выбран')}\n"
        f"Пользователь: {message.from_user.full_name} (@{message.from_user.username})\n"
        f"ID: {message.from_user.id}\n\n"
        f"Исходные данные пользователя: {data.get('device_raw_input', '')}\n"
    )
    if data.get("device_order"):
        admin_text += f"Распознанный номер заказа: {data.get('device_order')}\n"
    if data.get("device_phone"):
        admin_text += f"Распознанный номер телефона из заказа: {data.get('device_phone')}\n"
    admin_text += f"Контакт для связи: {contact}"
    delivered = await send_to_admin(message.bot, admin_text)
    await state.clear()
    await message.answer(
        (
            "Спасибо! Контакт получили, передали запрос администратору. Скоро свяжемся с вами."
            if delivered
            else "Контакт получили, но не удалось передать запрос администратору. "
            "Пожалуйста, проверьте корректность ADMIN_IDS."
        ),
        reply_markup=main_menu_keyboard(),
    )


async def handle_yes_no_available(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    data = await state.get_data()
    if callback.data == "yes_available":
        admin_text = (
            "📦 Запрос по устройству в сервисе\n\n"
            f"Офис: {data.get('chosen_office', 'Не выбран')}\n"
            f"Пользователь: {callback.from_user.full_name} (@{callback.from_user.username})\n"
            f"ID: {callback.from_user.id}\n\n"
            f"Исходные данные пользователя: {data.get('device_raw_input', '')}\n"
            f"Телефон для связи (доступен): {data.get('device_phone')}"
        )
        delivered = await send_to_admin(callback.bot, admin_text)
        await state.clear()
        await callback.message.edit_reply_markup(reply_markup=None)
        await callback.message.answer(
            (
                "Спасибо! Мы получили запрос и свяжемся с вами по указанному номеру."
                if delivered
                else "Номер получили, но не удалось передать запрос администратору. "
                "Проверьте корректность ADMIN_IDS."
            ),
            reply_markup=main_menu_keyboard(),
        )
        return
    await state.set_state(UserStates.waiting_contact_for_device)
    await callback.message.edit_text(
        cfg_text("ask_contact_text", "Пожалуйста, отправьте телефон или @username"),
        reply_markup=back_keyboard("back_device_info"),
    )


async def back_to_main_menu_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    await state.clear()
    await callback.message.edit_text(
        cfg_text("menu_text", "Выберите раздел:"),
        reply_markup=main_menu_keyboard(),
    )


async def back_repair_categories_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    data = await state.get_data()
    catalog = data.get("repair_catalog") or load_services_catalog()
    await state.set_state(UserStates.repair_choose_category)
    await callback.message.edit_text(
        "Выберите категорию устройства:",
        reply_markup=repair_categories_keyboard(catalog.get("categories", [])),
    )


async def back_repair_models_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    data = await state.get_data()
    models = data.get("repair_models", [])
    if not models:
        await back_repair_categories_handler(callback, state)
        return
    await state.set_state(UserStates.repair_choose_model)
    await callback.message.edit_text(
        "Выберите модель:",
        reply_markup=repair_models_keyboard(models, "back_repair_categories"),
    )


async def back_repair_problems_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    data = await state.get_data()
    services = data.get("repair_services", [])
    await state.set_state(UserStates.repair_choose_problem)
    await callback.message.edit_text(
        "Выберите услугу:",
        reply_markup=repair_services_keyboard(services, "back_repair_models"),
    )


async def back_repair_description_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    await state.set_state(UserStates.repair_description)
    await callback.message.edit_text(
        cfg_text("repair_description_prompt", "Опишите проблему подробнее."),
        reply_markup=back_keyboard("back_repair_problems"),
    )


async def back_repair_office_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    await state.set_state(UserStates.repair_choose_office)
    await callback.message.edit_text(
        cfg_text("repair_choose_office_text", "Выберите офис:"),
        reply_markup=offices_keyboard("repair_office", "back_repair_description"),
    )


async def back_repair_day_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    await state.set_state(UserStates.repair_choose_day)
    await callback.message.edit_text(
        cfg_text("repair_choose_day_text", "Выберите день записи:"),
        reply_markup=repair_day_keyboard(),
    )


async def back_repair_time_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    await state.set_state(UserStates.repair_choose_time)
    await callback.message.edit_text(
        cfg_text("repair_choose_time_text", "Выберите время:"),
        reply_markup=repair_time_keyboard(),
    )


async def back_repair_name_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    await state.set_state(UserStates.repair_ask_name)
    await callback.message.edit_text(
        cfg_text("repair_ask_name_text", "Как вас зовут?"),
        reply_markup=back_keyboard("back_repair_time"),
    )


async def back_device_office_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    await state.set_state(UserStates.waiting_device_office)
    await callback.message.edit_text(
        "Выберите офис, в который обращались:",
        reply_markup=offices_keyboard("device_office", "back_main_menu"),
    )


async def back_device_info_handler(callback: CallbackQuery, state: FSMContext) -> None:
    await callback.answer()
    await state.set_state(UserStates.waiting_device_info)
    await callback.message.edit_text(
        cfg_text("device_in_service_prompt", "Напишите номер заказа или телефон, указанный в заказе."),
        reply_markup=back_keyboard("back_device_office"),
    )


async def admin_command_handler(message: Message) -> None:
    if not is_admin(message.from_user.id):
        return
    await message.answer(
        "Админ-меню. Выберите категорию:",
        reply_markup=admin_menu_keyboard(),
    )


async def admin_root_handler(callback: CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer()
    await callback.message.edit_text(
        "Админ-меню. Выберите категорию:",
        reply_markup=admin_menu_keyboard(),
    )


async def admin_section_handler(callback: CallbackQuery) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer()
    section_key = callback.data.replace("admin_section_", "", 1)
    if section_key not in ADMIN_FIELDS_BY_SECTION:
        await callback.message.answer("Неизвестный раздел.")
        return
    await callback.message.edit_text(
        f"Категория: {ADMIN_SECTIONS.get(section_key, section_key)}\n\nВыберите параметр для редактирования:",
        reply_markup=admin_section_keyboard(section_key),
    )


async def admin_edit_button_handler(callback: CallbackQuery, state: FSMContext) -> None:
    if not is_admin(callback.from_user.id):
        await callback.answer("Нет доступа", show_alert=True)
        return
    await callback.answer()
    field_key = callback.data.replace("admin_edit_", "", 1)
    if field_key not in ADMIN_FIELD_LABELS:
        await callback.message.answer("Неизвестный параметр.")
        return
    if field_key == "excel_catalog_download":
        path = ensure_services_workbook()
        await callback.message.answer_document(
            FSInputFile(path),
            caption="Актуальный Excel-каталог услуг.",
        )
        await callback.message.answer(
            "Действия с Excel-каталогом:",
            reply_markup=admin_section_keyboard("json"),
        )
        await state.clear()
        return
    if field_key == "excel_catalog_upload":
        await state.set_state(AdminEditStates.waiting_excel_upload)
        await callback.message.edit_text(
            "Пришлите новый Excel-файл `.xlsx` с листом `services` и колонками: "
            "Категория, Модель, Услуга, Цена.",
            reply_markup=back_keyboard("admin_section_json"),
        )
        return

    get_cfg()
    await state.set_state(AdminEditStates.waiting_new_value)
    await state.update_data(edit_field=field_key)
    current_value = config.get(field_key)
    if field_key in ADMIN_JSON_FIELDS:
        serialized = json.dumps(current_value, ensure_ascii=False, indent=2)
        hint = "Отправьте новый JSON одним сообщением."
    else:
        serialized = str(current_value or "")
        hint = cfg_text("texts_edit_hint", "Отправьте новый текст одним сообщением.")

    await callback.message.edit_text(
        f"Редактирование: {ADMIN_FIELD_LABELS[field_key]}\n\n{hint}\n\nТекущее значение:\n{serialized}"
    )


async def admin_new_value_handler(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return

    data = await state.get_data()
    field_key = data.get("edit_field")
    if not field_key:
        await state.clear()
        await message.answer("Поле не выбрано.")
        return
    get_cfg()
    raw = message.text or ""
    if field_key in ADMIN_JSON_FIELDS:
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError:
            await message.answer("Ошибка JSON. Проверьте формат и отправьте снова.")
            return
        config[field_key] = parsed
    else:
        config[field_key] = raw
    save_config(config)
    await state.clear()
    await message.answer("Сохранено.", reply_markup=admin_menu_keyboard())


async def admin_excel_upload_handler(message: Message, state: FSMContext) -> None:
    if not is_admin(message.from_user.id):
        return
    if not message.document:
        await message.answer("Пришлите именно Excel-файл `.xlsx`.")
        return
    filename = message.document.file_name or ""
    if not filename.lower().endswith(".xlsx"):
        await message.answer("Нужен файл формата `.xlsx`.")
        return

    final_path = get_services_xlsx_path()
    target_dir = os.path.dirname(final_path) or "."
    os.makedirs(target_dir, exist_ok=True)

    with tempfile.NamedTemporaryFile(
        delete=False, suffix=".xlsx", dir=target_dir
    ) as tmp_file:
        tmp_path = tmp_file.name

    try:
        await message.bot.download(message.document, destination=tmp_path)
        validate_services_workbook(tmp_path)
        os.replace(tmp_path, final_path)
    except Exception as e:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        await message.answer(
            f"Не удалось загрузить Excel-каталог: {e}",
            reply_markup=admin_section_keyboard("json"),
        )
        return

    await state.clear()
    await message.answer(
        "Excel-каталог успешно обновлён.",
        reply_markup=admin_section_keyboard("json"),
    )


async def main() -> None:
    load_dotenv()
    bot_token = os.getenv("BOT_TOKEN")
    if not bot_token:
        raise RuntimeError("BOT_TOKEN is not set in environment")

    bot = Bot(
        token=bot_token,
        default=DefaultBotProperties(parse_mode=ParseMode.HTML),
    )
    dp = Dispatcher(storage=MemoryStorage())

    dp.message.register(start_handler, CommandStart())
    dp.message.register(admin_command_handler, Command("admin"))

    dp.callback_query.register(menu_repair_handler, F.data == "menu_repair")
    dp.callback_query.register(menu_hydro_handler, F.data == "menu_hydro")
    dp.callback_query.register(menu_price_handler, F.data == "menu_price")
    dp.callback_query.register(menu_device_service_handler, F.data == "menu_device_service")
    dp.callback_query.register(menu_warranty_handler, F.data == "menu_warranty")
    dp.callback_query.register(back_to_main_menu_handler, F.data == "back_main_menu")

    dp.callback_query.register(repair_category_handler, F.data.startswith("repair_cat_"))
    dp.callback_query.register(repair_model_handler, F.data.startswith("repair_model_"))
    dp.callback_query.register(repair_problem_handler, F.data.startswith("repair_service_"))
    dp.message.register(repair_other_problem_handler, UserStates.repair_other_problem)
    dp.message.register(repair_description_handler, UserStates.repair_description)
    dp.callback_query.register(repair_office_handler, F.data.startswith("repair_office_"))
    dp.callback_query.register(repair_day_handler, F.data.in_(["repair_day_tomorrow", "repair_day_other"]))
    dp.message.register(repair_date_input_handler, UserStates.repair_enter_date)
    dp.callback_query.register(repair_time_handler, F.data.startswith("repair_time_"))
    dp.message.register(repair_name_handler, UserStates.repair_ask_name)
    dp.message.register(repair_contact_handler, UserStates.repair_ask_contact)
    dp.callback_query.register(back_repair_categories_handler, F.data == "back_repair_categories")
    dp.callback_query.register(back_repair_models_handler, F.data == "back_repair_models")
    dp.callback_query.register(back_repair_problems_handler, F.data == "back_repair_problems")
    dp.callback_query.register(back_repair_description_handler, F.data == "back_repair_description")
    dp.callback_query.register(back_repair_office_handler, F.data == "back_repair_office")
    dp.callback_query.register(back_repair_day_handler, F.data == "back_repair_day")
    dp.callback_query.register(back_repair_time_handler, F.data == "back_repair_time")
    dp.callback_query.register(back_repair_name_handler, F.data == "back_repair_name")

    dp.callback_query.register(device_office_handler, F.data.startswith("device_office_"))
    dp.message.register(handle_device_info, UserStates.waiting_device_info)
    dp.message.register(handle_device_contact, UserStates.waiting_contact_for_device)
    dp.callback_query.register(handle_yes_no_available, F.data.in_(["yes_available", "no_available"]))
    dp.callback_query.register(back_device_office_handler, F.data == "back_device_office")
    dp.callback_query.register(back_device_info_handler, F.data == "back_device_info")

    dp.callback_query.register(admin_root_handler, F.data == "admin_root")
    dp.callback_query.register(admin_section_handler, F.data.startswith("admin_section_"))
    dp.callback_query.register(admin_edit_button_handler, F.data.startswith("admin_edit_"))
    dp.message.register(admin_new_value_handler, AdminEditStates.waiting_new_value)
    dp.message.register(admin_excel_upload_handler, AdminEditStates.waiting_excel_upload)

    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())

